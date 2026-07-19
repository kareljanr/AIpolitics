import csv
import re
from pathlib import Path

import openpyxl

path = Path(__file__).with_name("fps_taxex.xlsx")
wb = openpyxl.load_workbook(path, data_only=True)
rows_out = []

for sheet in wb.sheetnames:
    ws = wb[sheet]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    year_cols = [(i, h) for i, h in enumerate(header) if isinstance(h, int)]
    if not year_cols:
        print("no years", sheet)
        continue
    latest_i, latest_y = max(year_cols, key=lambda x: x[1])
    print(sheet, "latest year", latest_y)

    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        name = str(r[0]).replace("\xa0", " ").strip()
        val = r[latest_i] if latest_i < len(r) else None
        if not isinstance(val, (int, float)):
            continue
        ttype = r[1] if len(r) > 1 else None
        legal = r[2] if len(r) > 2 else None
        if sheet == "EIWT":
            legal = r[1] if len(r) > 1 else None
            ttype = "EIWT measure"
        if ttype is None and legal is None:
            continue
        # skip pure section titles
        if re.match(r"^\d+(\.\d+)*\.?\s*(Tax|Employment|Sectoral|Social|Various)", name, re.I):
            if not ttype or str(ttype).strip() in ("", "None"):
                continue
        rows_out.append(
            {
                "sheet": sheet,
                "name": name,
                "type": str(ttype).replace("\xa0", " ") if ttype else "",
                "legal": str(legal).replace("\xa0", " ") if legal else "",
                "year": latest_y,
                "amount_mio": float(val),
            }
        )

rows_out.sort(key=lambda x: -x["amount_mio"])
print("total measures", len(rows_out))
print("TOP 25:")
for i, r in enumerate(rows_out[:25], 1):
    print(f"{i:2} {r['amount_mio']:10.2f} mEUR  [{r['sheet']}] {r['name'][:90]}")

outp = Path(__file__).with_name("fps_taxex_parsed_latest.csv")
with outp.open("w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=["sheet", "name", "type", "legal", "year", "amount_mio"])
    w.writeheader()
    w.writerows(rows_out)
print("wrote", outp)
