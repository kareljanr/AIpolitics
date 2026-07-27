import openpyxl, csv, re
from pathlib import Path

wb = openpyxl.load_workbook(r"docs/doge/data/raw/fps_taxex.xlsx", data_only=True)

existing = []
with open(r"docs/doge/data/tax_expenditures.csv", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        try:
            ea = float(row["amount_eur"]) if row["amount_eur"] not in ("", "Unknown", None) else 0
        except Exception:
            ea = 0
        existing.append((row["taxex_id"], row["name"], ea))


def latest_amount(headers, row):
    best = None
    year = None
    for j, h in enumerate(headers):
        if h and str(h).isdigit() and j < len(row) and isinstance(row[j], (int, float)) and row[j] is not None:
            best = float(row[j])
            year = int(h)
    return year, best


def is_matched(m):
    nl = m["name"].lower()
    for eid, en, ea in existing:
        if ea and abs(ea - m["eur"]) / max(abs(m["eur"]), 1) < 0.03:
            return True
        if nl[:30] in en.lower() or en.lower()[:30] in nl:
            return True
        # keyword heuristics for already-imported families
        keys = [
            ("dtr", "dtr"),
            ("basic necessities", "vat_basic"),
            ("capital gains on shares eligible", "cit_cg"),
            ("pensions e.a", "pensions"),
            ("real estate (construction", "vat_realestate"),
            ("previous losses", "losses"),
            ("heating fuel", "heatoil"),
            ("innovation", "innovation"),
            ("horeca", "horeca"),
            ("foreign origin", "foreign"),
            ("health insurance", "health"),
            ("night work", "night"),
            ("nightshift", "night"),
            ("housing bonus", "housing"),
            ("job bonus", "jobbonus"),
            ("pension saving", "pension_save"),
            ("investment allowance", "invest"),
            ("researchers", "researchers"),
            ("small enterprises", "sme"),
            ("professional diesel", "pro_diesel"),
            ("company car", "company_cars"),
            ("fuel card", "fuel_cards"),
            ("charging", "charging"),
            ("meal voucher", "meal"),
            ("structural reduction", "structural"),
            ("shift work", "shift"),
            ("overtime", "overtime"),
            ("continuous work", "continuous"),
            ("kerosene", "kerosene"),
            ("social tariff", "social"),
            ("reduced rate", "gas_reduced"),  # careful
            ("unrealized capital gains", "unreal"),
        ]
        for kn, ke in keys:
            if kn in nl and ke in eid.lower():
                return True
    return False


allm = []
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c) if c is not None else "" for c in rows[0]]
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        if not name or name.lower().startswith("total"):
            continue
        if "measure in" in name.lower():
            continue
        y, amt = latest_amount(headers, row)
        if amt is None or abs(amt) < 1:
            continue
        allm.append(
            {
                "sheet": sn,
                "name": name,
                "year": y,
                "m_eur": amt,
                "eur": amt * 1e6,
                "type": str(row[1]).strip() if len(row) > 1 and row[1] else "",
                "legal": str(row[2]).strip() if len(row) > 2 and row[2] else "",
            }
        )

allm.sort(key=lambda x: -abs(x["eur"]))
print("TOP 55 latest-year measures:")
new_list = []
for i, m in enumerate(allm[:55]):
    matched = is_matched(m)
    flag = "HAVE" if matched else "NEW"
    print(f"{i+1:2} {flag:4} {m['sheet']:4} {m['year']} {m['m_eur']:10.2f}m | {m['name'][:75]}")
    if not matched:
        new_list.append(m)

print("\n=== NEXT NEW TOP 20 ===")
for i, m in enumerate(new_list[:20]):
    print(f"{i+1:2} {m['sheet']:4} {m['year']} {m['m_eur']:10.2f}m | {m['name'][:90]} | type={m['type'][:40]}")

# write candidates for CSV
out = Path("docs/doge/data/raw/tick137_taxex_new20.csv")
with out.open("w", encoding="utf-8", newline="") as f:
    w = csv.DictWriter(f, fieldnames=["sheet", "name", "year", "m_eur", "eur", "type", "legal"])
    w.writeheader()
    for m in new_list[:25]:
        w.writerow(m)
print("wrote", out)
