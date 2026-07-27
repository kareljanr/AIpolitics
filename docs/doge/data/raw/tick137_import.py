"""Import next 20 largest FPS tax expenditures not yet in tax_expenditures.csv."""
import openpyxl, csv, re
from pathlib import Path

wb = openpyxl.load_workbook(r"docs/doge/data/raw/fps_taxex.xlsx", data_only=True)

with open(r"docs/doge/data/tax_expenditures.csv", encoding="utf-8") as f:
    existing_rows = list(csv.DictReader(f))
    fieldnames = list(existing_rows[0].keys()) if existing_rows else [
        "taxex_id", "name", "level", "year", "amount_eur", "tax_type", "source_id", "confidence", "absurdity_seed", "notes"
    ]

existing_text = " ".join(
    (r["taxex_id"] + " " + r["name"]).lower() for r in existing_rows
)


def latest_amount(headers, row):
    best = None
    year = None
    for j, h in enumerate(headers):
        if h and str(h).isdigit() and j < len(row) and isinstance(row[j], (int, float)) and row[j] is not None:
            best = float(row[j])
            year = int(h)
    return year, best


# Known imported measure name fragments (strict)
HAVE_FRAGMENTS = [
    "dtr deduction",
    "basic necessities",
    "capital gains on shares eligible for fdi",
    "pensions e.a",
    "real estate (construction",
    "deduction previous losses",
    "gas oil low sulfur content - used as heating",
    "deduction for innovation",
    "horeca",
    "reduction taxes for income of foreign origin",
    "regional housing bonus",
    "federal housing bonus",
    "health insurance benefits",
    "night work",
    "nightshift work",
    "expressed unrealized capital gains",
    "job bonus",
    "pension saving",
    "investment allowance",
    "researchers employed",
    "reduced rates for small enterprises",
    "professional diesel",
    "company car",
    "fuel card",
    "charging card",
    "meal voucher",
    "structural reduction",
    "shift work",
    "overtime pay standard",  # EIWT only
    "continuous work",
    "kerosene",
    "social tariff",
    "reimbursement commuting",
    "bike",
    "electric car",
    "ecocheque",
    "stookolie",
    "heating gas oil",
    "aardgas",
    "lpg heating",
    "coal",
    "vat reduced gas",
    "package",
    "eiwt package",
    "fed total",
    "ffs",
]


def already_have(name: str) -> bool:
    nl = name.lower()
    for frag in HAVE_FRAGMENTS:
        if frag in nl:
            return True
    # id-ish tokens
    tokens = re.findall(r"[a-z0-9]{5,}", nl)
    for t in tokens:
        if t in existing_text and t not in ("income", "amount", "allowance", "deduction", "reduction", "exemption", "federal", "regional", "private", "companies"):
            # weak — skip
            pass
    return False


allm = []
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c) if c is not None else "" for c in rows[0]]
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        if not name or name.lower().startswith("total") or "measure in" in name.lower():
            continue
        y, amt = latest_amount(headers, row)
        if amt is None or abs(amt) < 50:  # >=50m EUR
            continue
        allm.append(
            {
                "sheet": sn,
                "name": name,
                "year": y,
                "m_eur": amt,
                "eur": int(round(amt * 1e6)),
                "type": str(row[1]).strip() if len(row) > 1 and row[1] else "",
            }
        )

allm.sort(key=lambda x: -abs(x["eur"]))
new = [m for m in allm if not already_have(m["name"])]
print(f"Candidates NEW >=50m: {len(new)}")
for i, m in enumerate(new[:25]):
    print(f"{i+1:2} {m['sheet']:4} {m['year']} {m['m_eur']:10.2f}m | {m['name'][:85]}")


def slug(s, max_len=40):
    s = re.sub(r"[^a-z0-9]+", "_", s.lower()).strip("_")
    return s[:max_len]


tax_type_map = {
    "PIT": "PIT",
    "CIT": "CIT",
    "WT": "WT",
    "EIWT": "EIWT",
    "VAT": "VAT",
    "EXC": "EXC",
    "NRT": "NRT",
}

# absurdity seeds: structural progressive/social design lower; opaque preferential higher
def absurdity(name, sheet, m_eur):
    nl = name.lower()
    if any(k in nl for k in ["tax free sum: basic", "dependent children", "birth allowance", "family allowance", "disability"]):
        return 2  # structural social/progressivity
    if "professional expenses" in nl:
        return 4  # large structural labour TE, debateable
    if "distinct taxation" in nl or "marital quotient" in nl:
        return 5
    if "venture capital" in nl:
        return 6
    if "unemployment" in nl:
        return 3
    if "long-term savings" in nl or "flexi" in nl:
        return 5
    if "withholding tax" in nl and "movable" in nl:
        return 5
    if "intra-group" in nl or "coordination" in nl:
        return 6
    if "refundable tax credit research" in nl:
        return 4
    return 4


picked = new[:20]
out_rows = []
for m in picked:
    tid = f"tx_{m['sheet'].lower()}_{slug(m['name'])}_{m['year']}"
    # shorten id
    tid = re.sub(r"_+", "_", tid)[:70]
    abs_s = absurdity(m["name"], m["sheet"], m["m_eur"])
    notes = f"FPS inventory {m['sheet']} sheet; {m['m_eur']:.2f} mEUR latest {m['year']}; type={m['type'] or 'n/a'}; tick137 next-20 import"
    if m["year"] and m["year"] < 2020:
        notes += "; HISTORICAL peak year not current regime"
    out_rows.append(
        {
            "taxex_id": tid,
            "name": f"{m['sheet']}: {m['name']}"[:120],
            "level": "federal",
            "year": str(m["year"]),
            "amount_eur": str(m["eur"]),
            "tax_type": tax_type_map.get(m["sheet"], m["sheet"]),
            "source_id": "src_fps_taxex_xlsx",
            "confidence": "strong",
            "absurdity_seed": str(abs_s),
            "notes": notes,
        }
    )

# append to tax_expenditures.csv
path = Path("docs/doge/data/tax_expenditures.csv")
with path.open("a", encoding="utf-8", newline="") as f:
    w = csv.DictWriter(f, fieldnames=fieldnames, lineterminator="\n")
    for r in out_rows:
        # only fields that exist
        w.writerow({k: r.get(k, "") for k in fieldnames})

print(f"\nAppended {len(out_rows)} rows")
for r in out_rows:
    print(r["taxex_id"], r["amount_eur"], r["name"][:70])
